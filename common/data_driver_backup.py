#coding:utf-8
import xlrd
import openpyxl
import configparser
import os
import re
# from xlutils.copy import copy
# import xlutils
# import json

class readexcel(object):
    def __init__(self,path,row):
        self.path = path
        self.row = row
        self.row =self.getrows()

    def getsheet(self):
        #通过索引获取一个工作表
        data = xlrd.open_workbook(self.path)
        sheet = data.sheet_by_index(0)
        return sheet

    #获取表的行数
    def getrows(self):
        # print(self.row)
        if self.row < 2:  # 如果self.row小于2时，则按excel中实际数据行数请求接口
            row = self.getsheet().nrows
        else:
            row = self.row  # # # # # #如果self.row大于等于2时，则按传入的self.row的值取数据
        return row

    #获取列数
    def getcol(self):
        col = self.getsheet().ncols
        return col

    # 判断各标题所在的列
    def getallcol(self):
        testallcol = []
        x = self.getcol()
        for i in range(0,x):
            testallcol.append(self.getsheet().cell(0, i).value)
        # print(TestUname0)
        return testallcol

    #分别获取每一列的数值
    def getuid(self):
        testuid = [] #存放第一列数据
        x = self.getallcol().index('Id')
        for i in range(1,self.row):
            # testuid.append(trim(self.getsheet().cell(i,x).value))
            testuid.append((self.getsheet().cell(i, x).value).strip())
        return testuid

    def getname(self):
        testname =[]
        x = self.getallcol().index('Name')
        for i in range(1,self.row):
            # testname.append(trim(self.getsheet().cell(i,x).value))
            testname.append((self.getsheet().cell(i, x).value).strip())
        return testname

    def config(self):
        # path_config = os.path.dirname(os.getcwd()) + '\config\config.ini'
        path_config = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'config', 'config.ini')
        print(path_config)
        cf = configparser.ConfigParser()
        cf.read(path_config)  # 读取配置
        print('sections:', cf, type(cf))
        secs = cf.sections()  # 读取sections
        print('sections:', secs, type(secs))
        opts = cf.options("HOST")  # 获取db section下的 options，返回list 如下，每个list元素为键值
        print('HOST:', opts, type(opts))
        kvs = cf.items("HOST")  # 获取db section 下的所有键值对，返回list 如下，每个list元素为键值对元组
        print(kvs)
        db_host = cf.get("HOST", "host")
        print(db_host)

    def geturl(self):
        testurl =[]#存放url
        # path_config = os.path.dirname(os.getcwd()) + '\config\config.ini'
        path_config = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'config', 'config.ini')
        print(path_config)
        cf = configparser.ConfigParser()
        cf.read(path_config)  # 读取配置
        # print('sections:', cf, type(cf))
        # secs = cf.sections()  # 读取sections
        # print('sections:', secs, type(secs))
        # opts = cf.options("HOST")  # 获取db section下的 options，返回list 如下，每个list元素为键值
        # print('HOST:', opts, type(opts))
        # kvs = cf.items("HOST")  # 获取db section 下的所有键值对，返回list 如下，每个list元素为键值对元组
        # print(kvs)
        db_host = cf.get("HOST", "host") #拼接域名
        print(db_host)
        print(testurl)
        # x = self.getallcol().index('Url')
        y = self.getallcol().index('Path')
        re0 = re.compile(r"http(.*)")  # 取Excel正则参数来校验

        for i in range(1,self.row):
            search1 = re0.search(self.getsheet().cell(i, y).value)
            if search1: # match1 = re0.match(zz1)   #match方法，类似searchtr法，re.match只匹配字符串的开始，如果字符串开始不符合正则表达式，则匹配失败，函数返回None；而re.search匹配整个字符串，直到找到一个匹配。
                # testurl.append(trim(self.getsheet().cell(i, y).value))
                testurl.append((self.getsheet().cell(i, y).value).strip())
                # print("打印url字段中的值的类型")
                # print(type(TestUrl),TestUrl)
            else:
                # testurl.append(db_host + trim(self.getsheet().cell(i,y).value))
                testurl.append(db_host + (self.getsheet().cell(i, y).value).strip())
        return testurl

    def getmethod(self):
        testmethod = [] #存放接口类型
        x = self.getallcol().index('Method')
        for i in range(1,self.row):
            # testmethod.append(trim(self.getsheet().cell(i,x).value))
            testmethod.append((self.getsheet().cell(i, x).value).strip())
        return testmethod

    # def getData(self):
    #
    #      TestData = [] #存放接口传送数据
    #      for i in range(1, self.row):
    #
    #          TestData.append(self.getSheet().cell(i,x).value)
    #      return TestData

    def getdata(self):
        testdata = []  # 存放接口传送数据
        x = self.getallcol().index('Data')
        for i in range(1, self.row):
            # testdata.append(trim(self.getsheet().cell(i, x).value))
            testdata.append((self.getsheet().cell(i, x).value).strip())
        return testdata
    def getopenid(self):
        testopenid = []  # 存放接口传送数据
        x = self.getallcol().index('OpenId')
        for i in range(1, self.row):
            # testopenid.append(trim(self.getsheet().cell(i, x).value))
            testopenid.append((self.getsheet().cell(i, x).value).strip())
        return testopenid

    def getbeforedata(self):
        testbeforedata = []  # 存放接口传送数据
        x = self.getallcol().index('BeforeData')
        for i in range(1, self.row):
            # testbeforedata.append(trim(self.getsheet().cell(i, x).value))
            testbeforedata.append((self.getsheet().cell(i, x).value).strip())
        return testbeforedata
    def getbeforesql(self):
        testbeforesql = []  # 存放接口传送数据
        x = self.getallcol().index('BeforeSql')
        for i in range(1, self.row):
            # testbeforesql.append(trim(self.getsheet().cell(i, x).value))
            testbeforesql.append((self.getsheet().cell(i, x).value).strip())
        return testbeforesql

    def getresql(self):
        testresql = []  # 存放接口传送数据
        x = self.getallcol().index('CheckSql')
        for i in range(1, self.row):
            # testresql.append(trim(self.getsheet().cell(i, x).value))
            testresql.append((self.getsheet().cell(i, x).value).strip())
        return testresql

    # def getstatuscode(self):
    #     teststatuscode = []# 存放status
    #     x = self.getallcol().index('Status')
    #     for i in range(1,self.row):
    #         teststatuscode.append(self.getsheet().cell(i,x).value)
    #     return teststatuscode


    def getrejson(self):
        testrejson = [] # 用列表存放正则校验json串
        x = self.getallcol().index('ExpectResult')
        for i in range(1,self.row):
            # testrejson.append(trim(self.getsheet().cell(i,x).value))
            testrejson.append((self.getsheet().cell(i, x).value).strip())
        return testrejson

    def get_paidanjigou(self):
        x = self.getallcol().index('派单机构')
        items = self.getsheet().col_values(x)
        return items

    def get_paidanjigou_yuan(self):
        x = self.getallcol().index('原派单机构')
        items = self.getsheet().col_values(x)
        return items

    def get_paidanjigou_xin(self):
        x = self.getallcol().index('新派单机构')
        items = self.getsheet().col_values(x)
        return items

    def get_crm_id(self):
        x = self.getallcol().index('CRM_ID')
        items = self.getsheet().col_values(x)
        return items

    def get_guanjia(self):
        x = self.getallcol().index('医美管家姓名')
        items = self.getsheet().col_values(x)
        return items

    def get_paidanyisheng(self):
        x = self.getallcol().index('派单医生')
        items = self.getsheet().col_values(x)
        return items

    def get_chengjiaojigou(self):
        x = self.getallcol().index('成交机构')
        items = self.getsheet().col_values(x)
        return items

    def get_chengjiaoyisheng(self):
        x = self.getallcol().index('成交医生')
        items = self.getsheet().col_values(x)
        return items

    # def writeresults(self): #copy一个表，然后新建一个表保存结果
    #     data = xlrd.open_workbook(self.path)
    #     book = copy(data)
    #     return book


    def openpyxl_read_sheet(self):
        book = openpyxl.load_workbook(self.path)
        # book = copy(data)
        # sheet =book["Sheet1"]
        return book

# def trim(s):
#     if s[:1] != ' ' and s[-1:] != ' ':
#         return s
#     elif s[:1] == ' ':
#         return trim(s[1:])
#     else:
#         return trim(s[:-1])

     
    
     

    
            
 
            
    
