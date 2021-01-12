#coding:utf-8
import xlrd
import openpyxl
import configparser
import os
import re
from xlutils.copy import copy
# import xlutils
import json

class readexcel(object):
    def __init__(self,path):
        self.path = path

    def getsheet(self):
        #通过索引获取一个工作表
        data = openpyxl.load_workbook(self.path)
        # print(data.worksheets)
        # openpyxl.load_workbook()
        # sheet = data.get_sheet_by_name(data.get_sheet_names()[0])
        sheet = data["Sheet1"]
        # print(sheet)
        # sheet = data.active()
        return sheet

    #获取表的行数
    def getrows(self):
        row = self.getsheet().max_row +1
        return row

    #获取列数
    def getcol(self):
        col = self.getsheet().max_column +1
        return col

    # 判断各标题所在的列
    def getallcol(self):
        testallcol = []
        x = self.getcol()
        # print(x)
        for i in range(1,x):
            testallcol.append(self.getsheet().cell(1, i).value)
        # print(TestUname0)
        return testallcol

    #分别获取每一列的数值
    def getuid(self):
        testuid = [] #存放第一列数据
        x = self.getallcol().index('Id')
        for i in range(2,self.getrows()):
            testuid.append(self.getsheet().cell(i,x).value)
        return testuid

    def getname(self):
        testname =[]
        x = self.getallcol().index('Name')
        for i in range(2,self.getrows()):
            testname.append(self.getsheet().cell(i,x).value)
        return testname

    def config(self):
        # path_config = os.path.dirname(os.getcwd()) + '\config\config.ini'
        path_config = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'config', 'config.ini')
        print(path_config)
        cf = configparser.ConfigParser()
        cf.read(path_config)  # 读取配置
        print('sections:', cf, type(cf))
        secs = cf.sections()  # 读取sections
        print('sections:', secs, type(secs))
        opts = cf.options("HOST")  # 获取db section下的 options，返回list 如下，每个list元素为键值
        print('HOST:', opts, type(opts))
        kvs = cf.items("HOST")  # 获取db section 下的所有键值对，返回list 如下，每个list元素为键值对元组
        print(kvs)
        db_host = cf.get("HOST", "host")
        print(db_host)

    def geturl(self):
        testurl =[]#存放url
        # path_config = os.path.dirname(os.getcwd()) + '\config\config.ini'
        path_config = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'config', 'config.ini')
        print(path_config)
        cf = configparser.ConfigParser()
        cf.read(path_config)  # 读取配置
        # print('sections:', cf, type(cf))
        # secs = cf.sections()  # 读取sections
        # print('sections:', secs, type(secs))
        # opts = cf.options("HOST")  # 获取db section下的 options，返回list 如下，每个list元素为键值
        # print('HOST:', opts, type(opts))
        # kvs = cf.items("HOST")  # 获取db section 下的所有键值对，返回list 如下，每个list元素为键值对元组
        # print(kvs)
        db_host = cf.get("HOST", "host") #拼接域名
        print(db_host)
        print(testurl)
        # x = self.getallcol().index('Url')
        y = self.getallcol().index('Path')
        re0 = re.compile(r"http(.*)")  # 取Excel正则参数来校验

        for i in range(2,self.getrows()):
            search1 = re0.search(str(self.getsheet().cell(i, y).value))
            if search1: # match1 = re0.match(zz1)   #match方法，类似searchtr法，re.match只匹配字符串的开始，如果字符串开始不符合正则表达式，则匹配失败，函数返回None；而re.search匹配整个字符串，直到找到一个匹配。
                testurl.append(self.getsheet().cell(i, y).value)
                # print("打印url字段中的值的类型")
                # print(type(TestUrl),TestUrl)
            else:
                testurl.append(db_host + str(self.getsheet().cell(i,y).value))
        return testurl

    def getmethod(self):
        testmethod = [] #存放接口类型
        x = self.getallcol().index('Method')
        for i in range(2,self.getrows()):
            testmethod.append(self.getsheet().cell(i,x).value)
        return testmethod

    # def getData(self):
    #
    #      TestData = [] #存放接口传送数据
    #      for i in range(1, self.getRows()):
    #
    #          TestData.append(self.getSheet().cell(i,x).value)
    #      return TestData

    def getdata(self):
        testdata = []  # 存放接口传送数据
        x = self.getallcol().index('Data')
        for i in range(2, self.getrows()):
            testdata.append(self.getsheet().cell(i, x).value)
        return testdata
    def getopenid(self):
        testopenid = []  # 存放接口传送数据
        x = self.getallcol().index('OpenId')
        for i in range(2, self.getrows()):
            testopenid.append(self.getsheet().cell(i, x).value)
        return testopenid

    def getbeforedata(self):
        testbeforedata = []  # 存放接口传送数据
        x = self.getallcol().index('BeforeData')
        for i in range(2, self.getrows()):
            testbeforedata.append(self.getsheet().cell(i, x).value)
        return testbeforedata
    def getbeforesql(self):
        testbeforesql = []  # 存放接口传送数据
        x = self.getallcol().index('BeforeSql')
        for i in range(2, self.getrows()):
            testbeforesql.append(self.getsheet().cell(i, x).value)
        return testbeforesql

    def getresql(self):
        testresql = []  # 存放接口传送数据
        x = self.getallcol().index('CheckSql')
        for i in range(2, self.getrows()):
            testresql.append(self.getsheet().cell(i, x).value)
        return testresql

    # def getstatuscode(self):
    #     teststatuscode = []# 存放status
    #     x = self.getallcol().index('Status')
    #     for i in range(1,self.getrows()):
    #         teststatuscode.append(self.getsheet().cell(i,x).value)
    #     return teststatuscode


    def getrejson(self):
        testrejson = [] # 用列表存放正则校验json串
        x = self.getallcol().index('ExpectResult')
        for i in range(2,self.getrows()):
            testrejson.append(self.getsheet().cell(i,x).value)
        return testrejson

    # def writeresults(self): #copy一个表，然后新建一个表保存结果
    #     data = xlrd.open_workbook(self.path)
    #     book = copy(data)
    #     return book
    def writeresults(self):
        data = openpyxl.load_workbook(self.path)
        book = copy(data)
        return book
        

     
    
     

    
            
 
            
    
